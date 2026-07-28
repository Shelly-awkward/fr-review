# -*- coding: utf-8 -*-
r"""
gen_inquiry_xlsx.py — inquiry.json → 「<年度>財務比率差異分析說明」查詢函 Excel（給受查公司填答）。

    python scripts/gen_inquiry_xlsx.py out/8304_114_inquiry.json out/8304_114_財務比率差異分析說明.xlsx

版面依過去實審實際樣本（110／112／114 年度三份「財務比率差異分析說明」）重排，六個 sheet：
  差異說明          查詢函本文：一 損益與成長率、二 同業比較、三 資產負債變動、四 產業發展、
                    五 風險事項、六 制式詢問十六項、七 資金貸與及背書保證、八 會計主管
                    （數字預填、「說明／ANS」欄留白給公司）
  基本資料          公司概況（XBRL 查得者預填，其餘標「行前查填」）
  財報資料          7 科目 × 近 6 年度
  財務比率          11 比率 × 近 6 年度（個別／上市櫃同業平均／所有同業平均）
  成長率            4 成長率 × 近 6 年度
  科目分析(內部)    全科目分析門檻與對應題號——內部覆核底稿，寄送公司前請刪除
同業平均無資料時留白並標註「請自公開資訊觀測站財務業務資訊貼入」——不得瞎編。
"""
import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill("solid", fgColor="DDEBF7")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
FONT = Font(name="微軟正黑體", size=11)
FONT_H = Font(name="微軟正黑體", size=11, bold=True)
FONT_T = Font(name="微軟正黑體", size=14, bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

FMT_AMT = "#,##0"
FMT_PCT = "0.00%"
FMT_NUM = "0.00"

# 差異說明「六、請說明以下事項」十六項制式詢問（文字照最新年度實審樣本）
Q16 = [
    "貴公司本期營運情形與所屬產業成長（或衰退）之變動是否相同及合理。",
    "本期較去年同期應收帳款及票據大幅增加，惟本期應付帳款及票據不增反減，有無重大異常。",
    "本期較去年同期應收帳款、存貨、其它應收款、預付款及轉投資大幅增加，有無重大異常。"
    "本期與關係人大幅背書保證或資金貸與，有無重大異常。",
    "本期較去年同期銷貨增加，惟本期機器設備無增加，且推銷費用不增反降，有無重大異常。",
    "本期應收關係人款未收回，是否已提列備抵呆帳，且是否與關係人間銷貨仍持續將增加，"
    "有無重大異常。",
    "本期與關係人間背書保證或資金貸與，有無重大異常。",
    "本期有無重大關係人交易。",
    "本期有無重大資產減損提列情形。",
    "本期有無認列資產減損迴轉金額，是否有重大異常。",
    "不動產、廠房及設備暨不動產之重大組成項目，及投資性不動產公允價值、方法及假設"
    "是否符合規定。",
    "自願改變會計政策或會計估計值變動中屬折舊(耗)性資產耐用年限、折舊（耗）方法與"
    "無形資產攤銷期間、攤銷方法之變動、殘值之變動及其公允價值之評價技術變動所致者，"
    "是否依證券發行人財務報告編製準則規定辦理、其他綜合損益組成項目變動是否合理。",
    "其它綜合損益組成項目變動是否合理。",
    "使用權資產及租賃負債是否依IFRS16規定認列及衡量，並揭露與租賃之攸關資訊，"
    "包括提供評估該租賃對公司財務狀況、財務績效與現金流量之影響及租賃活動之質性與"
    "量化相關資訊。",
    "金融資產是否依IFRS9規定認列及衡量，並依IFRS7規定揭露，包括金融資產對公司財務"
    "狀況與績效重要性之揭露資訊；金融資產所產生暴險之質性及量化資訊等。",
    "客戶合約之收入是否依IFRS15規定認列及衡量，並揭露與現金流量之性質、金額、時間及"
    "不確定性之綜合資訊，包括客戶合約之收入明細、合約餘額、履約義務、所作之重大判斷及"
    "判斷之改變，以及取得或履行客戶合約之成本中所認列之資產等。",
    "貴公司是否有對被投資公司持有具表決權之股份未超過50%且為被投資公司單一最大股東"
    "(縱使掌握之董事席次未過半)之情形?\n如有，貴公司是否已確實依IFRS 10第B38~B50段"
    "規定，評估對被投資公司是否具有權力。如貴公司評估未符合IFRS 10第7段所列3項控制力"
    "條件，僅具重大影響力之情形時，是否於財務報告附註揭露相關重大判斷過程及依據。",
]
CN_NUM = ["一", "二", "三", "四", "五", "六", "七", "八", "九", "十",
          "十一", "十二", "十三", "十四", "十五", "十六"]

Q7_YESNO = [  # (題文, 提示)
    ("{roc}年度是否從事資金貸與他人或背書保證事項:", ""),
    ("是否訂有作業程序：", "若有請提供"),
    ("所訂作業程序是否符合相關法令：", ""),
    ("是否依規定辦理個案評估:", ""),
    ("是否設立備查簿:", ""),
    ("是否業已依規定辦理資訊公開:", ""),
]
Q8 = [
    "貴公司之會計主管是否符合「發行人證券商證券交易所會計主管資格條件及專業進修辦法」"
    "第3條規定，其資格為何。",
    "貴公司之會計主管是否未有「發行人證券商證券交易所會計主管資格條件及專業進修辦法」"
    "第4條所列消極資格之情事。",
    "貴公司之會計主管之進修時數是否符合「發行人證券商證券交易所會計主管資格條件及"
    "專業進修辦法」之相關規定。(請檢附進修時數證明文件)",
    "另請貴公司確認是否業依「發行人證券商證券交易所會計主管資格條件及專業進修辦法」"
    "第9條及第10條規定辦理相關事項之申報。",
]


def style(cell, header=False, wrap=True):
    cell.border = BORDER
    cell.font = FONT_H if header else FONT
    cell.alignment = CENTER if header else WRAP
    if header:
        cell.fill = HEAD_FILL


def put_row(ws, r, values, header=False):
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=r, column=c, value=v)
        style(cell, header)


def na(v, blank_note=""):
    """None → 留白（或標註）；數值照放。"""
    return blank_note if v is None else v


def num(cell, v, fmt=FMT_AMT, scale=1):
    """數值格：None 留白；百分比以小數存、以 % 格式顯示（scale=0.01）。"""
    if v is None:
        return
    cell.value = v * scale
    cell.number_format = fmt


# ---------------------------------------------------------------- 差異說明

def merged(ws, r, c1, c2, text, header=False, height=None):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    for c in range(c1, c2 + 1):
        style(ws.cell(row=r, column=c), header)
    ws.cell(row=r, column=c1, value=text)
    if height:
        ws.row_dimensions[r].height = height
    return r + 1


def sec_title(ws, r, text):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    cell = ws.cell(row=r, column=1, value=text)
    cell.font = FONT_H
    cell.alignment = WRAP
    ws.row_dimensions[r].height = max(20, 18 * (len(text) // 55 + 1))
    return r + 1


def diff_thead(ws, r, cur_label, peer_label, sub, note_label="說明"):
    """兩列表頭：項次｜項目｜本期｜前期(或同業)｜比較增減(E:F)｜說明。sub＝E:F 副標。"""
    for c, v in enumerate(["項次", "項目", cur_label, peer_label], 1):
        ws.merge_cells(start_row=r, start_column=c, end_row=r + 1, end_column=c)
        style(ws.cell(row=r, column=c), header=True)
        style(ws.cell(row=r + 1, column=c), header=True)
        ws.cell(row=r, column=c, value=v)
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
    for c in (5, 6):
        style(ws.cell(row=r, column=c), header=True)
    ws.cell(row=r, column=5, value="比較增減")
    if len(sub) == 2:
        for c, v in zip((5, 6), sub):
            style(ws.cell(row=r + 1, column=c), header=True)
            ws.cell(row=r + 1, column=c, value=v)
    else:
        ws.merge_cells(start_row=r + 1, start_column=5, end_row=r + 1, end_column=6)
        for c in (5, 6):
            style(ws.cell(row=r + 1, column=c), header=True)
        ws.cell(row=r + 1, column=5, value=sub[0])
    ws.merge_cells(start_row=r, start_column=7, end_row=r + 1, end_column=7)
    style(ws.cell(row=r, column=7), header=True)
    style(ws.cell(row=r + 1, column=7), header=True)
    ws.cell(row=r, column=7, value=note_label)
    return r + 2


def answer_row(ws, r, label="", height=48):
    """留白作答列：A 標籤（如 ANS:），B:G 合併留白。"""
    style(ws.cell(row=r, column=1))
    ws.cell(row=r, column=1, value=label)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    for c in range(2, 8):
        style(ws.cell(row=r, column=c))
    ws.row_dimensions[r].height = height
    return r + 1


def question_row(ws, r, label, text):
    """題目列：A 編號、B:G 合併題文。"""
    style(ws.cell(row=r, column=1))
    ws.cell(row=r, column=1, value=label)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    for c in range(2, 8):
        style(ws.cell(row=r, column=c))
    ws.cell(row=r, column=2, value=text)
    ws.row_dimensions[r].height = max(20, 16 * (len(text) // 42 + 1))
    return r + 1


def build_diff_sheet(wb, q):
    meta, roc = q["meta"], q["meta"]["roc_year"]
    diff = q["diff"]
    company = meta.get("company") or meta.get("co")
    ws = wb.active
    ws.title = "差異說明"
    for col, w in zip("ABCDEFG", [7, 22, 15, 15, 13, 10, 58]):
        ws.column_dimensions[col].width = w

    title = f"{company}{roc}年度財務報告說明"
    ws.cell(1, 1, title).font = FONT_T
    ws.merge_cells("A1:G1")
    ws.cell(1, 1).alignment = Alignment(horizontal="center")
    ws.cell(2, 1, "單位:千元").font = FONT
    r = 3

    # 一、損益金額＋成長率兩期比較
    r = sec_title(ws, r, "一、下列財報資料請充分說明變動原因及合理性。")
    r = diff_thead(ws, r, f"{roc}年度", f"{roc-1}年度", ("金額", "％"))
    for it in diff["sec1_amounts"]:
        put_row(ws, r, [it["項次"], it["項目"], None, None, None, None, it["說明"]])
        num(ws.cell(r, 3), it["本期"])
        num(ws.cell(r, 4), it["前期"])
        num(ws.cell(r, 5), it["增減"])
        num(ws.cell(r, 6), it["變動%"], FMT_PCT, 0.01)
        if it["說明"]:
            ws.row_dimensions[r].height = max(32, 15 * (len(it["說明"]) // 40 + 1))
        r += 1
    for i, nm in enumerate(["合併報表營業收入淨額", "合併報表應收款項淨額\n(含應收票據)",
                            "合併報表存貨淨額"], len(diff["sec1_amounts"]) + 1):
        put_row(ws, r, [i, nm, "", "", "", "",
                        "（貴公司如編製合併報表請填列）" if i == len(diff["sec1_amounts"]) + 1
                        else ""])
        r += 1
    r = diff_thead(ws, r, f"{roc}年度", f"{roc-1}年度", ("百分點%",))
    for it in diff["sec1_growth"]:
        put_row(ws, r, [it["項次"], it["項目"], None, None, None, None, it["說明"]])
        if it["增減單位"] == "百分點":       # 成長率列：本期/前期/增減皆為百分比
            num(ws.cell(r, 3), it["本期"], FMT_PCT, 0.01)
            num(ws.cell(r, 4), it["前期"], FMT_PCT, 0.01)
            num(ws.cell(r, 5), it["增減"], FMT_PCT, 0.01)
        else:                              # 週轉率列：本期/前期為次數、增減為相對%
            num(ws.cell(r, 3), it["本期"], FMT_NUM)
            num(ws.cell(r, 4), it["前期"], FMT_NUM)
            num(ws.cell(r, 5), it["增減"], FMT_PCT, 0.01)
        r += 1

    # 二、與上市櫃同業比較
    r = sec_title(ws, r, "二、下列財務比率與上市櫃同業相較，請充分說明差異原因及合理性。"
                  "如上市櫃同業平均數尚難與貴公司情形進行比較分析，貴公司可自行選擇較相近"
                  "之上市櫃同業進行比較，並提供所選公司名稱及相關數據充分說明差異原因及"
                  "合理性。")
    r = diff_thead(ws, r, f"{roc}年度", "上市櫃\n同業數據", ("百分點%",))
    for it in diff["sec2_peer"]:
        put_row(ws, r, [it["項次"], it["項目"], None, None, None, None, it["說明"]])
        turn = it["增減單位"] == "%"
        fmt, scale = (FMT_NUM, 1) if turn else (FMT_PCT, 0.01)
        num(ws.cell(r, 3), it["本期"], fmt, scale)
        num(ws.cell(r, 4), it["同業"], fmt, scale)
        num(ws.cell(r, 5), it["增減"], FMT_PCT, 0.01)
        if it["同業"] is None:
            ws.cell(r, 4).fill = NOTE_FILL
        r += 1

    # 三、資產負債重大變動
    r = sec_title(ws, r, "三、本期較去年同期變動之原因及合理性。")
    r = diff_thead(ws, r, f"{roc}年度", f"{roc-1}年度", ("金額", "％"),
                   "請說明變動原因及合理性")
    if diff["sec3_bs"]:
        for it in diff["sec3_bs"]:
            put_row(ws, r, [it["項次"], it["項目"], None, None, None, None, it["說明"]])
            num(ws.cell(r, 3), it["本期"])
            num(ws.cell(r, 4), it["前期"])
            num(ws.cell(r, 5), it["增減"])
            num(ws.cell(r, 6), it["變動%"], FMT_PCT, 0.01)
            r += 1
    else:
        r = merged(ws, r, 1, 7, "（本期資產負債科目均未達重大變動分析門檻。）")

    # 四、產業發展
    r = sec_title(ws, r, f"四、請簡述貴公司{roc}年度所屬產業之發展情形。")
    r = answer_row(ws, r, height=64)

    # 五、風險事項
    r = sec_title(ws, r, "五、公司之風險事項")
    if diff["sec5_risk"]:
        for it in diff["sec5_risk"]:
            r = question_row(ws, r, "", it["問題"])
            r = answer_row(ws, r, height=64)
    else:
        r = merged(ws, r, 1, 7, "（本期無占資產總額10%以上之單一重大資產科目；"
                   "風險事項由承辦行前綜合判斷補列。）")

    # 六、制式詢問十六項
    r = sec_title(ws, r, "六、請說明以下事項：(若無以下情形，亦請回復\"無此情事\")")
    for i, text in enumerate(Q16):
        r = question_row(ws, r, f"({CN_NUM[i]})", text)
        r = answer_row(ws, r, "ANS:", height=40)

    # 七、資金貸與及背書保證
    r = sec_title(ws, r, "七、請說明貴公司從事資金貸與及背書保證是否符合公開發行公司"
                  "資金貸與及背書保證處理準則規定(請提供貴公司資金貸與及背書保證作業程序)")
    r = diff_thead(ws, r, f"{roc}年度", f"{roc-1}年度", ("金額", "％"))
    for i, nm in enumerate(["資金貸與他人金額", "背書保證金額"], 1):
        v = q["diff"]["sec7_loans"][nm]
        put_row(ws, r, [i, nm, None, None, None, None, ""])
        num(ws.cell(r, 3), v["本期"])
        num(ws.cell(r, 4), v["前期"])
        if v["本期"] is not None and v["前期"] is not None:
            num(ws.cell(r, 5), v["本期"] - v["前期"])
        if v["前期"] is None:
            ws.cell(r, 4, "（行前查填）")
        r += 1
    for text, hint in Q7_YESNO:
        style(ws.cell(r, 1))
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        for c in (2, 3):
            style(ws.cell(row=r, column=c))
        ws.cell(r, 2, text.format(roc=q["meta"]["roc_year"]))
        style(ws.cell(r, 4))
        ws.cell(r, 4, "是/否")
        style(ws.cell(r, 5))
        ws.cell(r, 5, hint)
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
        for c in (6, 7):
            style(ws.cell(row=r, column=c))
        r += 1

    # 八、會計主管
    r = sec_title(ws, r, "八、另請說明貴公司會計主管資格條件及進修時數是否符合"
                  "<發行人證券商證券交易所會計主管資格條件及專業進修辦法>之規定。")
    for i, text in enumerate(Q8):
        r = question_row(ws, r, f"({CN_NUM[i]})", text)
        r = answer_row(ws, r, height=40)

    # 連絡人
    r += 1
    ws.cell(r, 2, "貴公司連絡人").font = FONT_H
    r += 1
    ws.cell(r, 2, "　姓名：").font = FONT
    ws.cell(r, 4, "電話：").font = FONT
    r += 1
    ws.cell(r, 2, "　E-mail：").font = FONT
    return ws


# ---------------------------------------------------------------- 其餘各表

def sheet_header(ws, company, co, label, unit):
    ws.cell(1, 1, f"{company}（{co}）").font = FONT_T
    ws.cell(2, 1, label).font = FONT_H
    ws.cell(3, 1, unit).font = FONT
    return 5


def build_basic_sheet(wb, q):
    meta, roc = q["meta"], q["meta"]["roc_year"]
    ws = wb.create_sheet("基本資料")
    for col, w in zip("ABCDE", [18, 48, 18, 18, 18]):
        ws.column_dimensions[col].width = w
    company = meta.get("company") or meta.get("co")
    ws.cell(1, 1, f"{company}（{meta.get('co')}）").font = FONT_T
    ws.cell(2, 1, "基本資料").font = FONT_H
    ws.cell(3, 1, f"基本資訊（資料年/季：{roc}/4）").font = FONT

    fill = "（行前請至公開資訊觀測站「公司基本資料查詢」查填）"
    capital = q.get("capital")
    rows = [
        ("主要經營業務", fill),
        ("實收資本額", f"{int(capital):,}千元" if capital else fill),
        ("成立日期", fill), ("公發日期", fill),
        ("董事長", fill), ("總經理", fill),
        ("主要產業類別", na(meta.get("industry"), fill)),
        ("上市、上櫃同業公司", "（行前請至個別資料庫查詢系統／公開資訊觀測站查填）"),
    ]
    r = 4
    for k, v in rows:
        put_row(ws, r, [k, v])
        r += 1
    r += 1
    ws.cell(r, 1, "查核資訊（自財報XBRL申報資料取得）").font = FONT_H
    r += 1
    for k, v in [("簽證會計師事務所", meta.get("audit_firm")),
                 ("簽證會計師", meta.get("cpa")),
                 ("查核意見", meta.get("opinion")),
                 ("查核報告日", meta.get("report_date"))]:
        put_row(ws, r, [k, na(v, "－")])
        r += 1
    r += 1
    ws.cell(r, 1, "聯絡資訊").font = FONT_H
    r += 1
    put_row(ws, r, ["", "緊急聯絡人", "會計", "稽核", "股務"], header=True)
    r += 1
    for k in ["姓名", "職稱", "電話", "E-Mail"]:
        put_row(ws, r, [k, "", "", "", ""])
        r += 1
    return ws


def build_fin_sheet(wb, q, years_desc):
    meta = q["meta"]
    ws = wb.create_sheet("財報資料")
    company = meta.get("company") or meta.get("co")
    r = sheet_header(ws, company, meta.get("co"), "財報資料", "(單位：千元)")
    put_row(ws, r, ["項目"] + [f"{yy}年度" for yy in years_desc], header=True)
    r += 1
    for item, by_year in q["fin_data"].items():
        put_row(ws, r, [item] + [None] * len(years_desc))
        for c, yy in enumerate(years_desc, 2):
            v = by_year.get(str(yy), by_year.get(yy))
            if v is None:
                ws.cell(r, c, "－")
            else:
                num(ws.cell(r, c), v)
        r += 1
    r += 1
    ws.cell(r, 1, "註：「－」為現有XBRL資料未涵蓋之年度，行前請至個別資料庫查詢系統／"
                  "公開資訊觀測站查填。").font = FONT
    ws.column_dimensions["A"].width = 22
    for c in range(2, len(years_desc) + 2):
        ws.column_dimensions[get_column_letter(c)].width = 14
    return ws


def build_ratio_sheet(wb, q, years_desc):
    meta, roc = q["meta"], q["meta"]["roc_year"]
    ws = wb.create_sheet("財務比率")
    company = meta.get("company") or meta.get("co")
    r = sheet_header(ws, company, meta.get("co"), "財務比率", "(單位：％，次)")

    listed = q["ratios"].get("上市櫃同業平均") or {}
    allpeer = q["ratios"].get("所有同業平均") or {}
    by_year = q["ratios_by_year"]
    names = list(next(iter(by_year.values())).keys())

    # 表頭兩列：年度（跨3欄）／個別、上市櫃同業平均、所有同業平均
    style(ws.cell(r, 1), header=True)
    style(ws.cell(r + 1, 1), header=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=1)
    ws.cell(r, 1, "比率項目")
    for j, yy in enumerate(years_desc):
        c0 = 2 + j * 3
        ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c0 + 2)
        for c in range(c0, c0 + 3):
            style(ws.cell(r, c), header=True)
            style(ws.cell(r + 1, c), header=True)
        ws.cell(r, c0, f"{yy}年度")
        for c, v in zip(range(c0, c0 + 3), ["個別", "上市櫃同業平均", "所有同業平均"]):
            ws.cell(r + 1, c, v)
    r += 2
    for nm in names:
        put_row(ws, r, [nm] + [None] * (len(years_desc) * 3))
        for j, yy in enumerate(years_desc):
            c0 = 2 + j * 3
            v = by_year.get(str(yy), by_year.get(yy, {})).get(nm)
            if v is None:
                ws.cell(r, c0, "－")
            else:
                num(ws.cell(r, c0), v, FMT_NUM)
            # 同業平均僅本年度且使用者有提供才填；其餘留白待貼
            if yy == roc:
                lv, av = listed.get(nm), allpeer.get(nm)
                if lv is not None:
                    num(ws.cell(r, c0 + 1), lv, FMT_NUM)
                else:
                    ws.cell(r, c0 + 1).fill = NOTE_FILL
                if av is not None:
                    num(ws.cell(r, c0 + 2), av, FMT_NUM)
                else:
                    ws.cell(r, c0 + 2).fill = NOTE_FILL
        r += 1
    r += 1
    note = ws.cell(r, 1, "註：同業平均空白（黃底）欄位請自公開資訊觀測站「財務業務資訊」"
                         "查詢貼入（本表不代填、不推估）。流動比率、負債比率、現金流量比率、"
                         "股東權益報酬率、速動比率依XBRL財報數字計算，口徑可能與個別資料庫"
                         "查詢系統略異。")
    note.font = FONT
    note.fill = NOTE_FILL
    ws.column_dimensions["A"].width = 18
    for c in range(2, len(years_desc) * 3 + 2):
        ws.column_dimensions[get_column_letter(c)].width = 13
    return ws


def build_growth_sheet(wb, q, years_desc):
    meta = q["meta"]
    ws = wb.create_sheet("成長率")
    company = meta.get("company") or meta.get("co")
    r = sheet_header(ws, company, meta.get("co"), "成長率",
                     "(單位：％；空白＝缺基期資料)")
    put_row(ws, r, ["項目"] + [f"{yy}年度" for yy in years_desc], header=True)
    r += 1
    for item, by_year in q["growth4"].items():
        put_row(ws, r, [item] + [None] * len(years_desc))
        for c, yy in enumerate(years_desc, 2):
            v = by_year.get(str(yy), by_year.get(yy))
            if v is None:
                ws.cell(r, c, "－")
            else:
                num(ws.cell(r, c), v, FMT_NUM)
        r += 1
    ws.column_dimensions["A"].width = 20
    for c in range(2, len(years_desc) + 2):
        ws.column_dimensions[get_column_letter(c)].width = 13
    return ws


def build_analysis_sheet(wb, q):
    meta, roc = q["meta"], q["meta"]["roc_year"]
    ws = wb.create_sheet("科目分析(內部)")
    company = meta.get("company") or meta.get("co")
    ws.cell(1, 1, f"{company}（{meta.get('co')}）{roc}年度　科目分析總表").font = FONT_T
    ws.cell(2, 1, "本表為內部覆核底稿：列出所有分析科目及其分析門檻；標「★達標」者已於"
                  "「差異說明」出題請公司說明，未達標者亦列示以資覆核完整性。"
                  "寄送公司前請刪除本工作表。單位：新臺幣千元；比率為 % 或次。").font = FONT
    ws.cell(2, 1).fill = NOTE_FILL
    cov = q.get("data_coverage") or {}
    if cov:
        c3 = ws.cell(3, 1, "資料涵蓋年度：" + "、".join(str(x) for x in cov.get("資料年度", []))
                     + "　｜　成長率兩期比較："
                     + ("資料完整" if cov.get("足以比較兩期成長率")
                        else f"⚠ {cov.get('說明', '')}"))
        c3.font = FONT
        if not cov.get("足以比較兩期成長率"):
            c3.fill = NOTE_FILL
    cols = ["類別", "項目", f"{roc}年度", f"{roc-1}年度", "增減", "變動%",
            "分析門檻", "是否達標", "對應題號", "備註"]
    put_row(ws, 4, cols, header=True)
    r = 5
    for item in q.get("analysis", []):
        put_row(ws, r, [na(item.get(c)) for c in cols])
        if item.get("是否達標", "").startswith("★"):
            for c in range(1, len(cols) + 1):
                ws.cell(row=r, column=c).fill = NOTE_FILL
        r += 1
    ws.freeze_panes = "A5"
    for col, w in zip("ABCDEFGHIJ", [14, 22, 15, 15, 14, 11, 30, 15, 12, 40]):
        ws.column_dimensions[col].width = w
    return ws


def main():
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(2)
    with open(sys.argv[1], encoding="utf-8") as f:
        q = json.load(f)
    roc = q["meta"]["roc_year"]
    years_desc = sorted((int(k) for k in next(iter(q["fin_data"].values()))),
                        reverse=True)

    wb = Workbook()
    build_diff_sheet(wb, q)
    build_basic_sheet(wb, q)
    build_fin_sheet(wb, q, years_desc)
    build_ratio_sheet(wb, q, years_desc)
    build_growth_sheet(wb, q, years_desc)
    build_analysis_sheet(wb, q)

    wb.save(sys.argv[2])
    n_q = len(q.get("questions", []))
    print(f"✔ {sys.argv[2]}（差異說明 一～八節；科目分析 {len(q.get('analysis', []))} 項、"
          f"出題對照 {n_q} 題）")


if __name__ == "__main__":
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    main()
